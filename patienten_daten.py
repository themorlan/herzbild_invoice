import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pathlib
from typing import Union
from datetime import date, datetime
import traceback
import readchar

SOURCE_FILE = "/Users/marlon/Resilio Sync/Dokumente/Arbeit/Sonstiges/HerzBild/Rechnungsliste_HerzBild.xlsx"

pathlib.Path("Adresslisten").mkdir(exist_ok=True)
file_list = [file.resolve() for file in pathlib.Path("Adresslisten").glob('**/*.[cC][sS][vV]')]
if len(file_list) > 1:
    raise ImportError(f"Zu viele Adresslisten im Ordner: {file_list}")


def read_csv(file_path: Union[str, pathlib.Path]) -> pd.DataFrame:
    def get_index_row(_file_path):
        with open(_file_path, encoding="ISO-8859-1", mode="r") as file:
            for idx, line in enumerate(file.readlines()):
                if line.startswith("Pat-Nr"):
                    return idx
    _df = pd.read_csv(file_path, encoding="ISO-8859-1", delimiter=";", index_col=False, skip_blank_lines=False, header=get_index_row(file_path), dtype='string')
    _df = _df.drop('Untersuchungen pro Patient', axis=1)
    _df['Geb. Datum'] = _df['Geb. Datum'].apply(pd.to_datetime, errors='coerce', dayfirst=True)
    return _df


def match_with_data(birthday: date, patient_id: str, address_data: pd.DataFrame) -> pd.Series:
    _match = address_data.loc[(address_data['Pat-Nr'] == patient_id) & (address_data['Geb. Datum'] == birthday)]
    return _match


def main():
    def get_column_index(column: str, header: tuple) -> int:
        return header.index(column)
    book = openpyxl.load_workbook(SOURCE_FILE)
    sheet: Worksheet = book['Rechnungsliste']
    sheet_values = list(sheet.values)
    _header_index = 0
    for index, _tuple in enumerate(sheet_values):
        if "Rechnungsnummer" in _tuple:
            _header_index = index
            break
    # Remove 2 header rows from values
    _header = sheet_values[_header_index]
    sheet_values = sheet_values[_header_index + 1:]
    address_data = read_csv(file_list[0])
    for _row_nr, row_tuple in enumerate(sheet_values, start=_header_index + 2):
        _pat_id = str(row_tuple[get_column_index('PatientenID', _header)])
        _geb_dat = datetime.strptime(row_tuple[get_column_index('Geburtsdatum', _header)], "%d.%m.%Y") if type(row_tuple[get_column_index('Geburtsdatum', _header)]) == str else row_tuple[get_column_index('Geburtsdatum', _header)]
        _match = match_with_data(_geb_dat, _pat_id, address_data)
        if _match.empty:
            continue
        sheet.cell(row=_row_nr, column=get_column_index('Vorname', _header) + 1, value=_match['Vorname'].item())
        sheet.cell(row=_row_nr, column=get_column_index('Nachname', _header) + 1, value=_match['Name'].item())
        sheet.cell(row=_row_nr, column=get_column_index('Fallnummer', _header) + 1, value=int(_match['Fallnummer'].item()))
        if not _match['Adresse'].isna().any():
            _adresse = _match['Adresse'].item().split(", ")
            sheet.cell(row=_row_nr, column=get_column_index('Strasse', _header) + 1, value=_adresse[0])
            sheet.cell(row=_row_nr, column=get_column_index('PLZ', _header) + 1, value=int(_adresse[1].split(" ", 1)[0]))
            sheet.cell(row=_row_nr, column=get_column_index('Wohnort', _header) + 1, value=_adresse[1].split(" ", 1)[1])
    book.save(SOURCE_FILE)


if __name__ == '__main__':
    try:
        main()
    except:
        print(traceback.format_exc())
    finally:
        print("Drücke eine beliebige Taste um das Programm zu beenden...")
        k = readchar.readchar()
